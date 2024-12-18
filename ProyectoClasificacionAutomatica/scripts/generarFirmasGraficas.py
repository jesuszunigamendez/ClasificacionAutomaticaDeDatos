import openpyxl
from PIL import Image
import os
import argparse
import sys
import shutil
from datetime import datetime

class codificarExcel:
    def __init__(self,*,origen,destino,columnaClasificadora,ancho,alto,modo="entrenar"):
        """
            Clase que permite codificar un archivo de excel en arreglos bidimensionales

            Argumentos:
                origen: Archivo fuente
                destino: Carpeta donde se guardaran
                columnaClasificadora: columna donde se tomaran las clases a clasificar
                ancho: ancho de las firmas en pixeles
                alto: alto de las firmas en pixeles
                modo: indica si las firmas son para entrenar o para clasificar

            Retorna:
        """
        self.archivoOrigen = origen
        #print(self.carpetaDestinoPrincipal)
        self.elMayor = 0
        self.elMenor = 0
        self.anchoImagenes = ancho
        self.altoImagenes = alto
        self.columnaClasificadora = columnaClasificadora
        if modo == "entrenar":
            self.carpetaDestinoPrincipal = destino + "/"
            #if os.path.exists(self.carpetaDestinoPrincipal):
            #    shutil.rmtree(self.carpetaDestinoPrincipal)            
            self.procesarArchivo()
        else:
            self.carpetaDestinoPrincipal = destino + "/"
            self.carpetaDestinoSecundaria = ""
            #if os.path.exists(self.carpetaDestinoPrincipal):
            #    shutil.rmtree(self.carpetaDestinoPrincipal)            
            self.generarImagenesDeArchivo()
        print("Archivo procesado correctamente")
        #self.guardarImagen(arrayRGB="",CarpetaDestino="",nombreimagen="")



    def generarImagenesDeArchivo(self):
        """
            Funcion que procesa todo el archivo generando firmas aptas para clasificar
        """
        #  se abre e; excel
        wb = openpyxl.load_workbook(self.archivoOrigen)
        # se selecciona la primera hoja
        hoja = wb.active
        #se inicializan variables
        numeroFila = 0
        numeroCelda = 0
        corrida = datetime.now().strftime("%d_%m_%y_%H_%M_%S")
        CarpetaDestino = "Firmas_Clasificar" + "_" + str(corrida) 
        self.carpetaDestinoSecundaria = CarpetaDestino
        #self.columnaClasificadora = 10
        # for que recorre fila for fila
        for fila in hoja.iter_rows(values_only=True):
            numeroFila = numeroFila + 1            
            print("procesando fila ",numeroFila)            
            numeroCelda = 0
            arrayRGB = []
            arrayceldas = []
            # for que recorre celda por celda
            contadorCelda = 0
            for celda in fila:
                contadorCelda = contadorCelda + 1
                numeroCelda = numeroCelda + 1
                # si son las demas filas entonces se procesa cada celda como una linea de pixeles para la imagen 
                if (numeroFila != 1):# and (identificadorCelda != 0):
                    if numeroCelda == self.columnaClasificadora:
                        a = 1
                        #print("excluida",celda)
                    else:
                        #if numeroCelda == self.columnaClasificadora
                        # todo el valor de la celda se convierte en un arreglo de valores unicode convertidos a decimal
                        arrayceldas.append(celda)
                        unicodedec, unicodehex = self.cadenaADecimalHexadecimal(str(celda))
                        # para cada valor decimal en el arreglo equivalenter a la informacion de la celda se obtiene su equivalente en RGB
                        for dec in unicodedec:
                            arrayRGB.append(self.array_binarios_a_RGB(self.dividirBinarioenBytes(self.decialABinario24(dec))))
                            # se actualizan baderas de valores
                            #################################
                            if dec > self.elMayor:
                                self.elMayor = dec
                            if dec < self.elMenor:
                                self.elMenor = dec
                            #################################                        
                        # al terminar de convertir los caracteres a su equivalente unicode y de unicode a su equivalente RGB se agrega un identificador de fin de celda    
                        arrayRGB.append([256,256,256])

            if numeroFila != 1:
                # cuando se procesa toda una fila se guarda la imagen resultante en la carpeta correspondiente
                nombre = os.path.basename(self.archivoOrigen) + "_fila_" + str(numeroFila) + ".png"
                arrayRGBAjustado = self.ajustarColores(arrayRGB)
                self.guardarImagen(arrayRGBAjustado,CarpetaDestino,nombre)



    def procesarArchivo(self):
        """
            Funcion que procesa todo el archivo generando firmas aptas para entrenar
        """
        #  se abre e; excel
        wb = openpyxl.load_workbook(self.archivoOrigen)
        # se selecciona la primera hoja
        hoja = wb.active
        #se inicializan variables
        numeroFila = 0
        numeroCelda = 0
        identificadorCelda = 0
        CarpetaDestino = "failFolder"
        # for que recorre fila for fila
        for fila in hoja.iter_rows(values_only=True):
            numeroFila = numeroFila + 1            
            print("procesando fila ",numeroFila)            
            numeroCelda = 0
            arrayRGB = []
            arrayceldas = []
            # for que recorre celda por celda
            contadorCelda = 0
            for celda in fila:
                contadorCelda = contadorCelda + 1
                numeroCelda = numeroCelda + 1
                # si es la primera fila solo se identifica cual va a ser la columna que indica la carpeta organizadora
                if numeroFila == 1:
                    #if celda == self.columnaClasificadora:
                    if contadorCelda == self.columnaClasificadora:
                        identificadorCelda = numeroCelda
                # si son las demas filas entonces se procesa cada celda como una linea de pixeles para la imagen 
                elif (numeroFila != 1) and (identificadorCelda != 0):
                #if (numeroFila != 1) and (identificadorCelda != 0):
                    # si estamos sobre la celda que indica la carpeta destino entonces indicamos el nombre que traiga la celda como carpeta destino
                    if identificadorCelda == numeroCelda:
                        CarpetaDestino = str(celda)
                    # si es otra celda entonces tomamos su informacion y la codificamos a valores RGB                        
                    else:
                        # todo el valor de la celda se convierte en un arreglo de valores unicode convertidos a decimal
                        arrayceldas.append(celda)
                        unicodedec, unicodehex = self.cadenaADecimalHexadecimal(str(celda))
                        # para cada valor decimal en el arreglo equivalenter a la informacion de la celda se obtiene su equivalente en RGB
                        for dec in unicodedec:
                            #print("aqui")
                            #print(dec)
                            #print(self.decialABinario24(dec))
                            #print(self.dividirBinarioenBytes(self.decialABinario24(dec)))
                            #print(self.array_binarios_a_RGB(self.dividirBinarioenBytes(self.decialABinario24(dec))))
                            #if celda == None or celda == "None":
                                #arrayRGB.append([256,256,256])
                                #break
                            #else:
                            arrayRGB.append(self.array_binarios_a_RGB(self.dividirBinarioenBytes(self.decialABinario24(dec))))
                            #print("fin")
                            # se actualizan baderas de valores
                            #################################
                            if dec > self.elMayor:
                                self.elMayor = dec
                            if dec < self.elMenor:
                                self.elMenor = dec
                            #################################                        
                        # al terminar de convertir los caracteres a su equivalente unicode y de unicode a su equivalente RGB se agrega un identificador de fin de celda    
                        arrayRGB.append([256,256,256])
            if identificadorCelda == 0:
                print("La columna " + self.columnaClasificadora + " no se encontro en el archivo a procesar")                        
                sys.exit()
            if numeroFila != 1:
                # cuando se procesa toda una fila se guarda la imagen resultante en la carpeta correspondiente
                nombre = os.path.basename(self.archivoOrigen) + "_fila_" + str(numeroFila) + ".png"
                arrayRGBAjustado = self.ajustarColores(arrayRGB)
                self.guardarImagen(arrayRGBAjustado,CarpetaDestino,nombre)



    def ajustarColores(self,arrayRGB):
        """
            Funcion que permite convertir un arreglo de valores en un nuevo arreglo que derscarta los valores
            nulos, recive un arreglo de la forma [[0,0,#],[0,#,#]...,[0#,0,#]] y devuelve [[#,#,#],...,[#,#,#]]

            Argumentos: 
                arrayRGB: arreglo que contiene valores de pixeles

            Retorna:
                arrayPixelesARetornar: arreglo que contiene los pixeles validos de array RGB
        """
        #primera accion, obtenemos todos los colores de los pixeles diferentes de cero y los guardamos en el array de colores validos
        # #######################################
        # print("imprimimos el array original")
        # for p in arrayRGB:
        #     if p != [256,256,256]:
        #         print(p)
        #     else:
        #         print("")
        # ######################################
        arrayColoresValidos = []
        for pixel in arrayRGB:
            if pixel != [256,256,256]:
                for color in pixel:
                    if color != 0:
                        arrayColoresValidos.append(color)
            else:
                arrayColoresValidos.append(256)
        # ################################################                
        # print("imprimimos colores validos")
        # print(arrayColoresValidos)
        # for p in arrayColoresValidos:
        #     if p != 256:
        #         print(p)
        #     else:
        #         print("")
        # #sys.exit()                
        # #######################################        
        arrayPixelesARetornar = []
        #segunda accion, colvemos a poner los colores en grupos de trews para generar un pixel
        tamanio = len(arrayColoresValidos)
        contadorColor = 0
        pixelOriginal = [255,255,255]
        pixel = pixelOriginal.copy()
        for i in range(tamanio - 1):
            color = arrayColoresValidos[i]
            colorSiguiente = arrayColoresValidos[i + 1]
            if color == 256:
                arrayPixelesARetornar.append([256,256,256])
            else:
                pixel[contadorColor] = color
                contadorColor = contadorColor + 1
                if (contadorColor == 3) or  (colorSiguiente == 256):
                    #print(pixel)
                    arrayPixelesARetornar.append(pixel.copy())
                    pixel = pixelOriginal.copy()
                    contadorColor = 0                
        # #######################################
        # print("imprimimos el array nuevo")
        # for p in arrayPixelesARetornar:
        #     if p != [256,256,256]:
        #         print(p)
        #     else:
        #         print("")
        # ######################################

        return arrayPixelesARetornar


    def guardarImagen(self,arrayRGB,CarpetaDestino,nombreimagen):
        """
            Funcion que permite guardar un arreglo de pixeles como una imagen

            Argumentos:
                arrayRGB: arreglo de pixeles
                CarpetaDestino: carpeta donde se fguardara la imagen
                nombreimagen: combre de la imagen a guardar
            Retorna 
        """
        #arrayRGB = [[256,256,256],[0,0,0],[0,0,0]]
        if len(arrayRGB) == 0:
            print("Error. El arreglo de pixeles para " + str(nombreimagen) + " esta vacio") 
        else:
            # se calcula el tamanio de la imagen que se generara
            ancho = 1
            alto = 1
            anchoMayor = 0
            for pixel in arrayRGB:
                if pixel == [256,256,256]:
                    alto = alto + 1
                    ancho = 1
                else:
                    if ancho > anchoMayor:
                        anchoMayor = ancho
                    ancho = ancho + 1                        
            ancho = anchoMayor

            if self.altoImagenes != 0:
                if self.altoImagenes > alto:
                    alto = self.altoImagenes
                else:
                    print("El alto en pixeles necesario para escribir la imagen " + nombreimagen + " es mayor a (" + str(self.altoImagenes) + ") por lo que se ajusto al tamaño necesario (" + str(alto) + ")")

            if self.anchoImagenes != 0:
                if self.anchoImagenes > ancho:
                    ancho = self.anchoImagenes
                else:
                    print("El ancho en pixeles necesario para escribir la imagen " + nombreimagen + " es mayor a (" + str(ancho) + ") por lo que se ajusto al tamaño necesario (" + str(anchoMayor) + ")")


            # se crea una nueva imagen en modo RGB
            try:
                imagen = Image.new("RGB", (ancho, alto), "white")  # Color de fondo blanco por defecto
            except Exception as e:
                print("Ocurrio un error al generar la firma grafica")
                print(e)
                return 0
            CarpetaDestino = self.carpetaDestinoPrincipal + CarpetaDestino
            #print(CarpetaDestino)
            # se asignan los valores RGB a cada píxel
            #print("Los valores rgb a guardarse como imagen en  ",CarpetaDestino)        
            x = 0
            y = 0
            contador = 0
            for pixel in arrayRGB:
                contador = contador + 1
                if pixel != [256,256,256]:
                    #print(pixel,end="")
                    try:
                        imagen.putpixel((x, y), tuple(pixel))
                    except Exception as e:
                        print(e)
                        print(pixel)
                        sys.exit()
                    x = x + 1
                else:
                    #print("")
                    y = y + 1
                    x = 0
            ## Primer píxel (fila 1, columna 1)
            #imagen.putpixel((1, 0), tuple(pixeles_rgb[1]))  # Segundo píxel (fila 1, columna 2)
            #imagen.putpixel((0, 1), tuple(pixeles_rgb[2]))  # Tercer píxel (fila 2, columna 1)
            # Crear la carpeta si no existe
            if not os.path.exists(CarpetaDestino):
                os.makedirs(CarpetaDestino)
    
            # Guardar la imagen en la carpeta correspondiente
            ruta_imagen = os.path.join(CarpetaDestino, nombreimagen)
            #print(ruta_imagen)
            imagen.save(ruta_imagen)
    
            #print(f"Imagen guardada en '{ruta_imagen}'")


            



    def cadenaADecimalHexadecimal(self,cadena):
        """
            Funcion que convierte una cadena de caracteres a sus valores decimales y hexadecimales codificados segu unicode

            Argumentos
                cadena: cadena de bits

            Retorna
                tupla que contiene la conversion
        """
        # Codificar la cadena en UTF-8
        utf8_bytes = cadena.encode('utf-8')

        # Convertir los bytes a valores decimales
        utf8_codificacion_decimal = [byte for byte in utf8_bytes]

        # Convertir los bytes a valores hexadecimales
        utf8_codificacion_hex = [f"{byte:02X}" for byte in utf8_bytes]

        return utf8_codificacion_decimal, utf8_codificacion_hex  


    def decialABinario24(self,decimal):
        """
            Funcion que convierte un valor decimal a bonario de 24 bits

            Argumentos:
                decimal: valor a convertir

            Retorna: valor convertido
        """
        if decimal >= 0:
            binario = bin(decimal)[2:] 
        else:
            binario = bin(decimal & 0xFFFFFF)[2:]

        binario_24bits = binario.zfill(24)

        return binario_24bits
    
    def dividirBinarioenBytes(self,binario):
        """
            Funcion que divide un valor binario de 24 bits en un arreglo de bytes

            Argumentos:
                binario: valor a convertir

            Retorna: arreglo de bytes 
        """
        #se procesa caracter por caracter 
        cadena = str(binario)
        #print(cadena)
        arrayBytes = []
        byte = ""
        contador = 0
        for caracter in cadena:
            byte = byte + caracter
            contador = contador + 1
            if contador == 8:
                arrayBytes.append(byte)
                byte = ""
                contador = 0
        return arrayBytes                

    def array_binarios_a_RGB(self,arrayBinario):
        """
            Funcion que convierte un arreglo de bytes en binario a sus valor respectiivos en decimal o RGB

            Argumentos:
                arrayBinario: arreglo a convertir

            Retorna: arreglo convertido
        """
        arrayDecimales = []
        for binario in arrayBinario:
            # Validar si la cadena es un número binario válido
            if not all(bit in '01' for bit in binario):
                raise ValueError("La cadena no es un número binario válido.")
    
            # Convertir binario a decimal
            decimal = 0
            longitud = len(binario)

            for i in range(longitud):
                # Convertir el bit a decimal y leerlo desde el final
                bit = int(binario[longitud - 1 - i])  
                decimal += bit * (2 ** i)

            arrayDecimales.append(decimal)
        


        return arrayDecimales




