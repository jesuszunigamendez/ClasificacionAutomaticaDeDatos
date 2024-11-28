import tkinter as tk
from tkinter import filedialog
import tensorflow as tf
import numpy as np
from tensorflow.keras.preprocessing import image
import json
import cv2
import os
from openpyxl import load_workbook
import platform
import re


class Clasificar:
    def __init__(self,*,rutaModelo,rutaArchivo,rutaDataSetClasificar,rutaResultado,ancho,alto):
        """
            Clase que permite clasificar las imagenes que se encuentran dentro de un directorio
        """
        self.rutaModelo = rutaModelo
        self.nombreModelo = self.rutaModelo.split('/')[-1]
        self.nombreModelo = os.path.splitext(self.nombreModelo)[0]
        self.model = tf.keras.models.load_model(self.rutaModelo)
        self.rutaArchivo = rutaArchivo
        self.rutaDataSetClasificar = rutaDataSetClasificar
        #se obtiene la lista de archivos para clasificacion
        self.archivosClasificar = [f for f in os.listdir(self.rutaDataSetClasificar) if f.endswith('.png')]
        self.rutaResultado = rutaResultado
        self.rutaClases = self.obtenerRutaClases()
        self.indices = self.obtenerIndices()
        self.arrayClasificar = []
        self.clasificar()
        rutaGuardada = self.guardarExcel()
        self.abrirExcel(rutaGuardada)

    def abrirExcel(self,ruta):
        """
            Funcion que permite abrir un archvio dada uuna ruta

            Argumentos:
                ruta: archivo a abrir
        """
        # Abrir el archivo en el sistema operativo
        if platform.system() == "Windows":
            os.startfile(ruta)  # En Windows
        elif platform.system() == "Darwin":
            os.system(f"open {ruta}")  # En macOS
        else:
            os.system(f"xdg-open {ruta}")  # En Linux

    def guardarExcel(self):
        """"
            Esta funcion guarda el excel que contiene las predicciones, usando el excel de procedencia de las firma 
            y modificandolo de tal forma que se tenga la prediccion y la certeza de la misam
        """
        # Obtener el nombre del archivo original sin la ruta completa
        nombreArchivo = os.path.basename(self.rutaArchivo)
        # Crear el nuevo nombre con el prefijo
        
        nombreGuardar = f"Resultados__usando_modelo_{self.nombreModelo}_Para_{nombreArchivo}"
        # Combinar la nueva ruta y el nuevo nombre
        rutaGuardado = os.path.join(self.rutaResultado, nombreGuardar)

        # Cargar el archivo Excel
        workbook = load_workbook(self.rutaArchivo)

        # Seleccionar la hoja activa (o especificar la hoja por nombre)
        hoja = workbook.active

        # Obtener el número actual de columnas
        num_columnas = hoja.max_column

        # Agregar encabezados en la primera fila de las nuevas columnas
        hoja.cell(row=1, column=num_columnas + 1, value="Resultado")
        hoja.cell(row=1, column=num_columnas + 2, value="Certeza %")

        # Recorrer las filas desde la segunda para agregar los datos
        contadorFila = 1
        for fila in hoja.iter_rows(min_row=2, max_col=num_columnas + 2):  # max_col para recorrer las nuevas columnas
            contadorFila = contadorFila + 1
            for arreglo in self.arrayClasificar:
                if contadorFila == int(arreglo[0]):
                    resultado = arreglo[1]
                    confianza = round((float(arreglo[2]) * 100) , 2)
                    break
            # Penúltima columna (nueva columna añadida)
            fila[num_columnas].value = str(resultado)
    
            # Última columna (nueva columna añadida)
            fila[num_columnas + 1].value = str(confianza)

        # Guardar los cambios en la nueva ubicación con el nuevo nombre
        workbook.save(rutaGuardado)

        print(f"Se guardó el archivo modificado como: {rutaGuardado}")
        return rutaGuardado






    # Función para clasificar la imagen
    def clasificadorImagen(self,model, img_path, class_indices):
        """
            Funcion que permite clasificar una imagen usando un modelo preentrenado
        """
        # Obtén la forma de entrada del modelo
        input_shape = model.input_shape[1:3]  # Esto es (img_shape, img_shape)
        img = self.cargarImagen(img_path, input_shape[0])
        prediction = model.predict(img)
        if class_indices:
            class_names = list(class_indices.keys())
            predicted_class = class_names[np.argmax(prediction)]
            confidence = np.max(prediction)
            return predicted_class, confidence
        else:
            return None, None
        
    # Función para cargar y preprocesar la imagen
    def cargarImagen(self, filename, img_shape):
        """
            Funcion que permite obtener las caracteristicas de la imagen para permitir normalizarla 
        """        
        img = image.load_img(filename, target_size=(img_shape, img_shape))
        img = image.img_to_array(img)
        img = np.expand_dims(img, axis=0)
        img = img / 255.0
        return img        
        
    def clasificar(self):
        """
            Funcion que permite clasificar las imagenes contenidas en un directorio
        """
        # se procesa imagen por  imagen
        for imagen in self.archivosClasificar:
            buscar = re.search(r"_(\d+)\.png$", imagen)
            # si la imagen existe en el direc
            if buscar:
                #el formato del nombre de las imagenes ermina como Fila# entonces e obtiene ese # para permitir ubicarlo en el excel
                numero = buscar.group(1) 
                rutaImagen = self.rutaDataSetClasificar + "/" + imagen
                #se obtiene la prediccion para esta imagen y la confianza de la misma
                prediccion, confianza = self.clasificadorImagen(self.model, rutaImagen, self.indices)
                if prediccion is not None and confianza is not None:
                    self.arrayClasificar.append([numero,prediccion,confianza])
                else:
                    self.arrayClasificar.append([numero,"Fail","0"])
            else:
                print("Error al procesar la firma " + imagen)     
    

    def obtenerIndices(self):
        """
            Funcion que permiute obtener los indices de un archivo JSON
        """
        indices = None
        try:
            with open(self.rutaClases, 'r') as f:
                indices = json.load(f)
                return indices
        except FileNotFoundError:
            return None
        
    def obtenerRutaClases(self):
        """
            Funcion que se encarga de generar la ruta de las clases con las que se clasificara
        """
        partes = self.rutaModelo.split("/")
        ruta = ""
        for parte in partes:
            if parte.endswith(".h5"):
                nombre = parte[:-2]
                ruta = ruta + nombre + "json"
            else:
                ruta = ruta + parte + "/"
        return ruta